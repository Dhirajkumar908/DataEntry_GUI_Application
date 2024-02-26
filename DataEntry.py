from tkinter import *
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

root=Tk()

root.geometry('550x400')
root.title("DataEntry In ExcelSheet")
root.config(bg="#E0EEE0")

#funcutions 


file=pathlib.Path('DataEntry_file.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Name"
    sheet["B1"]="age"
    sheet["C1"]="Contact No."
    sheet["D1"]="Gender"
    sheet["E1"]="Address"
    file.save("DataEntry_file.xlsx")

def submit():
    name= name1.get()
    age=age1.get()
    contect=contect1.get()
    gender=c4.get()
    address=en5.get(1.0, END)

    file=openpyxl.load_workbook("DataEntry_file.xlsx")
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=age)
    sheet.cell(column=3, row=sheet.max_row, value=contect)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)
    file.save(r'DataEntry_file.xlsx')

def clear():
    en1.delete(0, END)
    en2.delete(0,END)
    en3.delete(0, END)
    en5.delete(1.0, END)
    en1.focus()

#GUI 

Label(root, text="Enter Datails Here", font=('Arial, bold', 20), bg="#C1CDC1", width=40, pady=5).pack()

l1=Label(root, text="Name", font=('Arial,bold', 15), bg="#E0EEE0")
l1.place(x=20, y=60)

l2=Label(root, text="Age", font=('Arial,bold', 15), bg="#E0EEE0")
l2.place(x=20, y=110)

l3=Label(root, text="Contact No.", font=('Arial,bold', 15), bg="#E0EEE0")
l3.place(x=20,y=160)

l4=Label(root, text="Gender:",  font=('Arial,bold', 15), bg="#E0EEE0" )
l4.place(x=280, y=110)

l5 = Label(root, text="Address:", font=('Arial,bold', 15), bg="#E0EEE0")
l5.place(x=20, y=210)


name1=StringVar()
age1=StringVar()
contect1=StringVar()

en1=Entry(root, textvariable=name1, font=('Arial, bold', 15), width=30, bd=2)
en1.place(x=130, y=60)

en2=Entry(root, textvariable=age1, font=('Arial, bold',15), width=10, bd=2)
en2.place(x=130, y=110)

en3=Entry(root, textvariable=contect1, font=('Arial, bold',15), width=30, bd=2)
en3.place(x=130, y=160)

en5 = Text(root, width=41, height=5, bd=2)
en5.place(x=130, y=210)



c4=Combobox(root, values=["Male","Femel"], font=('Arial', 15), width=8)
c4.place(x=355, y=110)
c4.set("Male")

bt1=Button(root, text="Submit", font=('Arial,bold', 15), bg="#C1CDC1", bd=2, command=submit)
bt1.place(x=100, y=330)

bt2=Button(root, text="Clear", font=('Arial,bold', 15), bg="#C1CDC1", bd=2, command=clear)
bt2.place(x=200, y=330)

bt3=Button(root, text="Exit", font=('Arial,bold', 15), bg="#C1CDC1", bd=2, command=lambda :root.destroy())
bt3.place(x=300, y=330)

root.mainloop()